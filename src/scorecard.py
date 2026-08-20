import os
import sys
import numpy as np
import pandas as pd
import joblib
from sqlalchemy import create_engine
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to sys.path to allow importing from src
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from src.db import get_db_params

def calculate_credit_score(pd_val, w=4.03):
    """
    Scale Probability of Default (PD) into a FICO-like credit score (300 to 850).
    Score = Offset + Factor * ln(Odds)
    where Odds = w * (1 - PD) / PD
    """
    # Clip PD to avoid log of 0 or division by zero
    pd_val = np.clip(pd_val, 0.0001, 0.9999)
    odds = (1 - pd_val) / pd_val
    true_odds = w * odds
    
    # Calibration parameters
    # Base Odds of 4:1 (average default rate in portfolio is 20%) maps to 600 points
    pdo = 20
    base_odds = 4
    base_score = 600
    
    factor = pdo / np.log(2)
    offset = base_score - factor * np.log(base_odds)
    
    score = offset + factor * np.log(true_odds)
    # Clip between standard FICO limits
    return int(np.clip(score, 300, 850))

def generate_excel_scorecard():
    print("1. Loading final XGBoost model and preprocessor metadata...")
    model_path = os.path.join("output", "model.joblib")
    meta_path = os.path.join("data", "processed", "preprocessor_meta.joblib")
    
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Trained model not found at {model_path}. Please run src/train.py first.")
        
    xgb = joblib.load(model_path)
    preprocessor_meta = joblib.load(meta_path)
    w = float(preprocessor_meta.get("scale_pos_weight", 4.03))
    
    print("2. Connecting to PostgreSQL to pull original test loan metadata...")
    params = get_db_params()
    connection_uri = f"postgresql://{params['user']}:{params['password']}@{params['host']}:{params['port']}/{params['dbname']}"
    engine = create_engine(connection_uri)
    
    # Query original features for loans issued in 2017+ (test set)
    query = """
        SELECT loan_id, loan_amnt, annual_inc, grade, sub_grade, int_rate, purpose, default_flag 
        FROM loans 
        WHERE issue_year >= 2017;
    """
    test_raw = pd.read_sql_query(query, con=engine)
    print(f"   Loaded {test_raw.shape[0]:,} raw test rows from PostgreSQL.")
    
    # Load processed X_test features
    X_test = pd.read_csv(os.path.join("data", "processed", "X_test.csv"))
    meta_test = pd.read_csv(os.path.join("data", "processed", "meta_test.csv"))
    
    print("3. Generating predictions and scaling credit scores...")
    # Generate probability of default (PD)
    y_prob = xgb.predict_proba(X_test)[:, 1]
    
    # Add predictions to metadata
    meta_test['probability_of_default'] = y_prob
    # Calculate credit scores
    meta_test['credit_score'] = [calculate_credit_score(p, w) for p in y_prob]
    
    # Define Risk Tiers
    # Low Risk: Score >= 660 (PD <= 5.2%)
    # Medium Risk: Score 580-659 (PD 5.2% - 15.6%)
    # High Risk: Score < 580 (PD > 15.6%)
    def get_risk_tier(score):
        if score >= 660:
            return 'Low'
        elif score >= 580:
            return 'Medium'
        else:
            return 'High'
            
    meta_test['risk_tier'] = meta_test['credit_score'].apply(get_risk_tier)
    
    # Merge model scores back with original raw columns
    scored_df = pd.merge(test_raw, meta_test[['loan_id', 'probability_of_default', 'credit_score', 'risk_tier']], on='loan_id')
    
    # Calculate key aggregate summaries for dashboard
    print("4. Calculating aggregate summaries...")
    total_loans = len(scored_df)
    total_defaults = scored_df['default_flag'].sum()
    overall_default_rate = total_defaults / total_loans
    
    # Group by Risk Tier
    tier_summary = scored_df.groupby('risk_tier').agg(
        total_loans=('loan_id', 'count'),
        defaults=('default_flag', 'sum'),
        avg_score=('credit_score', 'mean'),
        avg_int_rate=('int_rate', 'mean')
    ).reset_index()
    tier_summary['default_rate'] = tier_summary['defaults'] / tier_summary['total_loans']
    
    # Re-order risk tiers logically: Low, Medium, High
    tier_order = {'Low': 0, 'Medium': 1, 'High': 2}
    tier_summary['sort_key'] = tier_summary['risk_tier'].map(tier_order)
    tier_summary = tier_summary.sort_values('sort_key').drop(columns=['sort_key'])
    
    # Cross-tab comparison: Lending Club Grade vs our Risk Tier (Default Rates)
    crosstab_count = pd.crosstab(scored_df['grade'], scored_df['risk_tier'], values=scored_df['loan_id'], aggfunc='count', margins=True).fillna(0)
    crosstab_defaults = pd.crosstab(scored_df['grade'], scored_df['risk_tier'], values=scored_df['default_flag'], aggfunc='sum', margins=True).fillna(0)
    
    # Reorder columns logically: Low, Medium, High, All
    cols_order = ['Low', 'Medium', 'High', 'All']
    # Use reindex to ensure all columns exist even if counts are 0, avoiding KeyErrors
    crosstab_count = crosstab_count.reindex(columns=cols_order, fill_value=0)
    crosstab_defaults = crosstab_defaults.reindex(columns=cols_order, fill_value=0)
    
    crosstab_rate = (crosstab_defaults / crosstab_count).fillna(0)
    
    print("5. Generating formatted Excel workbook via openpyxl...")
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    # ── Style Definitions ───────────────────────────────────────────────
    font_family = "Segoe UI"
    
    # Colors (Sleek professional palette)
    steel_blue = "2C5E8A"
    light_steel = "F2F6F9"
    dark_gray = "333333"
    border_color = "D9D9D9"
    
    # Fills
    header_fill = PatternFill(start_color=steel_blue, end_color=steel_blue, fill_type="solid")
    zebra_fill = PatternFill(start_color=light_steel, end_color=light_steel, fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Soft alert fills
    low_risk_fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid") # soft teal
    med_risk_fill = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid") # soft orange
    high_risk_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid") # soft red
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="2C5E8A")
    section_font = Font(name=font_family, size=12, bold=True, color="333333")
    header_font = Font(name=font_family, size=10, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=10, bold=True, color="333333")
    regular_font = Font(name=font_family, size=10, color="333333")
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border_side = Side(style='thin', color=border_color)
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom_border = Border(top=thin_border_side, bottom=Side(style='double', color='333333'))
    
    # ── Sheet 1: Executive Dashboard ────────────────────────────────────
    ws_dash = wb.create_sheet(title="Executive Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_dash['A1'] = "Credit Risk Scorecard Dashboard"
    ws_dash['A1'].font = title_font
    ws_dash.row_dimensions[1].height = 25
    
    # Metadata info
    ws_dash['A2'] = "Lending Club Portfolio Validation Set (2017 - 2018)"
    ws_dash['A2'].font = Font(name=font_family, size=10, italic=True)
    ws_dash.row_dimensions[2].height = 15
    
    # 1. Portfolio Key Metrics Summary
    ws_dash['A4'] = "Portfolio Overview Summary"
    ws_dash['A4'].font = section_font
    
    overview_headers = ["Metric Description", "Value", "Notes"]
    for col_num, h in enumerate(overview_headers, start=1):
        cell = ws_dash.cell(row=5, column=col_num)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align if col_num != 2 else right_align
    ws_dash.row_dimensions[5].height = 20
    
    metrics = [
        ("Total Active Loans Evaluated", total_loans, "All validation set records (2017-2018)"),
        ("Total Charged-off Defaults", total_defaults, "Actual loans defaults"),
        ("Portfolio Default Rate", overall_default_rate, "Historical baseline default rate"),
    ]
    
    for idx, (m_desc, val, note) in enumerate(metrics, start=6):
        ws_dash.cell(row=idx, column=1, value=m_desc).font = regular_font
        val_cell = ws_dash.cell(row=idx, column=2, value=val)
        val_cell.font = bold_font
        
        # Number formats
        if "%" in m_desc or "Rate" in m_desc:
            val_cell.number_format = "0.00%"
        else:
            val_cell.number_format = "#,##0"
            
        ws_dash.cell(row=idx, column=3, value=note).font = regular_font
        
        for col_num in range(1, 4):
            c = ws_dash.cell(row=idx, column=col_num)
            c.border = thin_border
            c.fill = zebra_fill if idx % 2 == 0 else white_fill
        ws_dash.row_dimensions[idx].height = 20
        
    # 2. Performance by Model Risk Tier
    ws_dash['A11'] = "Credit Performance by Model Risk Tier"
    ws_dash['A11'].font = section_font
    
    tier_headers = ["Model Risk Tier", "Total Loans", "Actual Defaults", "Model Score Range", "Default Rate (Actual)", "Avg Interest Rate"]
    for col_num, h in enumerate(tier_headers, start=1):
        cell = ws_dash.cell(row=12, column=col_num)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_num == 1 or col_num == 4 else (right_align if col_num != 1 else left_align)
    ws_dash.row_dimensions[12].height = 20
    
    ranges = {
        'Low': "Score 660 - 850",
        'Medium': "Score 580 - 659",
        'High': "Score 300 - 579"
    }
    
    for idx, row in tier_summary.iterrows():
        r_row = idx + 13
        tier = row['risk_tier']
        
        t_cell = ws_dash.cell(row=r_row, column=1, value=tier)
        t_cell.font = bold_font
        t_cell.alignment = center_align
        
        # Apply soft color code fills
        fill = low_risk_fill if tier == 'Low' else (med_risk_fill if tier == 'Medium' else high_risk_fill)
        t_cell.fill = fill
        
        l_cell = ws_dash.cell(row=r_row, column=2, value=row['total_loans'])
        l_cell.number_format = "#,##0"
        l_cell.font = regular_font
        
        d_cell = ws_dash.cell(row=r_row, column=3, value=row['defaults'])
        d_cell.number_format = "#,##0"
        d_cell.font = regular_font
        
        rng_cell = ws_dash.cell(row=r_row, column=4, value=ranges[tier])
        rng_cell.alignment = center_align
        rng_cell.font = regular_font
        
        rate_cell = ws_dash.cell(row=r_row, column=5, value=row['default_rate'])
        rate_cell.number_format = "0.00%"
        rate_cell.font = bold_font
        
        int_cell = ws_dash.cell(row=r_row, column=6, value=row['avg_int_rate']/100.0 if row['avg_int_rate'] > 1.0 else row['avg_int_rate'])
        int_cell.number_format = "0.00%"
        int_cell.font = regular_font
        
        for col_num in range(1, 7):
            c = ws_dash.cell(row=r_row, column=col_num)
            c.border = thin_border
            if col_num != 1:
                c.fill = zebra_fill if r_row % 2 == 0 else white_fill
        ws_dash.row_dimensions[r_row].height = 22
        
    # 3. Validation Analysis: Lending Club Grade vs Model Risk Tier
    ws_dash['A18'] = "Cross-Validation Matrix: Lending Club Grade vs. Model Risk Tier (Actual Default Rates)"
    ws_dash['A18'].font = section_font
    
    matrix_headers = ["Lending Club Grade", "Low Risk Tier", "Medium Risk Tier", "High Risk Tier", "All Portfolio"]
    for col_num, h in enumerate(matrix_headers, start=1):
        cell = ws_dash.cell(row=19, column=col_num)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align if col_num == 1 else right_align
    ws_dash.row_dimensions[19].height = 20
    
    grades = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'All']
    for idx, gr in enumerate(grades):
        r_row = idx + 20
        g_cell = ws_dash.cell(row=r_row, column=1, value=f"Grade {gr}" if gr != 'All' else "Total Portfolio")
        g_cell.font = bold_font
        g_cell.alignment = left_align
        
        for c_idx, tier_name in enumerate(['Low', 'Medium', 'High', 'All'], start=2):
            rate = crosstab_rate.loc[gr, tier_name]
            loans_count = crosstab_count.loc[gr, tier_name]
            
            c = ws_dash.cell(row=r_row, column=c_idx)
            c.font = regular_font
            
            if loans_count > 0:
                c.value = rate
                c.number_format = "0.00%"
                # Apply soft alert highlights on default rates inside the table cells
                if tier_name != 'All' and gr != 'All':
                    if rate >= 0.25:
                        c.fill = high_risk_fill
                        c.font = Font(name=font_family, size=10, bold=True, color="C0392B")
                    elif rate >= 0.10:
                        c.fill = med_risk_fill
                        c.font = Font(name=font_family, size=10, bold=True, color="D68910")
                    else:
                        c.fill = low_risk_fill
                        c.font = Font(name=font_family, size=10, color="117A65")
            else:
                c.value = "-"
                c.alignment = right_align
                
            if gr == 'All' or tier_name == 'All':
                c.font = bold_font
                c.border = double_bottom_border
            else:
                c.border = thin_border
                
            if (gr == 'All' or tier_name == 'All') and not c.fill.fill_type:
                c.fill = zebra_fill
                
        ws_dash.row_dimensions[r_row].height = 20
        
    # Auto-adjust column widths for Dashboard
    for col in ws_dash.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_dash.column_dimensions[col_letter].width = max(max_len + 4, 12)
    ws_dash.column_dimensions['A'].width = 35
    ws_dash.column_dimensions['D'].width = 20
    
    # ── Sheet 2: Account Level Details (Subset) ──────────────────────────
    ws_details = wb.create_sheet(title="Account Level Scoring")
    ws_details.views.sheetView[0].showGridLines = True
    
    detail_headers = [
        "Loan ID", "Loan Amount ($)", "Annual Income ($)", "Lending Club Grade", 
        "Interest Rate (%)", "Model PD (%)", "Model Credit Score", "Assigned Risk Tier", "Default Status"
    ]
    
    for col_num, h in enumerate(detail_headers, start=1):
        cell = ws_details.cell(row=1, column=col_num)
        cell.value = h
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_num in [1, 4, 7, 8, 9] else (right_align if col_num in [2, 3, 5, 6] else left_align)
    ws_details.row_dimensions[1].height = 22
    
    # Load a representative sample of test data to keep the Excel responsive (max 50,000 accounts)
    sample_limit = 50000
    if len(scored_df) > sample_limit:
        print(f"   Test set size is {len(scored_df):,}. Sampling {sample_limit:,} rows for Sheet 2...")
        details_sample = scored_df.sample(sample_limit, random_state=42).copy()
    else:
        details_sample = scored_df.copy()
        
    # Sort details by Credit Score descending
    details_sample = details_sample.sort_values('credit_score', ascending=False)
    
    print("   Writing scored account data rows...")
    row_idx = 2
    for idx, r in details_sample.iterrows():
        # Formatted inputs
        ws_details.cell(row=row_idx, column=1, value=r['loan_id']).alignment = center_align
        
        lamnt = ws_details.cell(row=row_idx, column=2, value=r['loan_amnt'])
        lamnt.number_format = "$#,##0"
        
        inc = ws_details.cell(row=row_idx, column=3, value=r['annual_inc'])
        inc.number_format = "$#,##0"
        
        gr_cell = ws_details.cell(row=row_idx, column=4, value=r['grade'])
        gr_cell.alignment = center_align
        
        int_rate = ws_details.cell(row=row_idx, column=5, value=r['int_rate']/100.0 if r['int_rate'] > 1.0 else r['int_rate'])
        int_rate.number_format = "0.00%"
        
        pd_val = ws_details.cell(row=row_idx, column=6, value=r['probability_of_default'])
        pd_val.number_format = "0.00%"
        
        sc_cell = ws_details.cell(row=row_idx, column=7, value=r['credit_score'])
        sc_cell.alignment = center_align
        sc_cell.font = bold_font
        
        tier = r['risk_tier']
        t_cell = ws_details.cell(row=row_idx, column=8, value=tier)
        t_cell.alignment = center_align
        t_cell.font = bold_font
        # Soft risk fill color
        t_cell.fill = low_risk_fill if tier == 'Low' else (med_risk_fill if tier == 'Medium' else high_risk_fill)
        
        def_flag = r['default_flag']
        d_cell = ws_details.cell(row=row_idx, column=9, value="DEFAULT" if def_flag == 1 else "PERFORMING")
        d_cell.alignment = center_align
        d_cell.font = bold_font
        # soft red/green text alerts for default status
        d_cell.fill = high_risk_fill if def_flag == 1 else low_risk_fill
        d_cell.font = Font(name=font_family, size=10, bold=True, color="C0392B" if def_flag == 1 else "117A65")
        
        # Apply default border/zebra stripe
        for col_num in range(1, 10):
            c = ws_details.cell(row=row_idx, column=col_num)
            c.border = thin_border
            if col_num != 8 and col_num != 9:
                c.fill = zebra_fill if row_idx % 2 == 0 else white_fill
                c.font = regular_font
                
        ws_details.row_dimensions[row_idx].height = 18
        row_idx += 1
        
    # Auto-adjust column widths for Details sheet
    for col in ws_details.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_details.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # Save the Excel Workbook
    output_path = os.path.join("output", "scorecard_report.xlsx")
    wb.save(output_path)
    print(f"6. Excel scorecard report generated successfully and saved to {output_path}")

if __name__ == '__main__':
    try:
        generate_excel_scorecard()
    except Exception as e:
        print(f"Error generating scorecard: {e}")
